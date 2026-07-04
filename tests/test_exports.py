"""Tests για τα exports (/api/exports): print, μαζική εκτύπωση, XLSX, ICS.

Πριν από αυτό το αρχείο τα exports δεν είχαν καμία κάλυψη — τα πρώτα tests
εδώ κλειδώνουν την υπάρχουσα συμπεριφορά (single print/ICS) και τα υπόλοιπα
οδηγούν τα νέα features: /print?all=teachers|classes (ένα έγγραφο με
page-break ανά καθηγητή/τμήμα) και /xlsx (workbook με φύλλο ανά οντότητα).
"""
from __future__ import annotations

import io

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.database import Base, get_db
from backend.models import (
    Classroom,
    Lesson,
    Period,
    SchoolClass,
    SchoolSettings,
    Student,
    StudentClassEnrollment,
    Subject,
    Teacher,
    TimetableSlot,
    TimetableSolution,
)
from backend.routers import exports as exports_router


@pytest.fixture()
def client():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine)
    s = Session()

    s.add(SchoolSettings(school_name="T", days_per_week=5, institution_type="frontistirio"))
    math = Subject(name="Μαθηματικά", short_name="ΜΑΘ", color="#000")
    physics = Subject(name="Φυσική", short_name="ΦΥΣ", color="#000")
    t1 = Teacher(name="Νικολάου Μαρία", short_name="ΝΜ", color="#000")
    t2 = Teacher(name="Παπαδόπουλος Γιώργος", short_name="ΠΓ", color="#000")
    c1 = SchoolClass(name="Α1 Λυκείου", short_name="Α1")
    c2 = SchoolClass(name="Β1 Λυκείου", short_name="Β1")
    room1 = Classroom(name="Αίθουσα 1", short_name="Α1θ")
    room2 = Classroom(name="Αίθουσα 2", short_name="Α2θ")
    periods = [
        Period(name=f"{i}η", short_name=str(i), start_time=f"{15+i}:00",
               end_time=f"{15+i}:50", is_break=False, sort_order=i)
        for i in range(1, 4)
    ]
    student = Student(first_name="Νίκη", last_name="Κοντού")
    s.add_all([math, physics, t1, t2, c1, c2, room1, room2, student, *periods])
    s.commit()
    for o in [math, physics, t1, t2, c1, c2, room1, room2, student, *periods]:
        s.refresh(o)

    l1 = Lesson(subject_id=math.id, teacher_id=t1.id, class_id=c1.id,
                periods_per_week=2, duration=1)
    l2 = Lesson(subject_id=physics.id, teacher_id=t2.id, class_id=c2.id,
                periods_per_week=1, duration=1)
    s.add_all([l1, l2])
    s.commit()
    s.refresh(l1)
    s.refresh(l2)
    s.add(StudentClassEnrollment(student_id=student.id, class_id=c1.id))

    sol = TimetableSolution(name="Λύση Α", status="optimal")
    s.add(sol)
    s.commit()
    s.refresh(sol)
    # t1/Μαθηματικά: Δευτέρα 1η + Τρίτη 2η (Αίθουσα 1) · t2/Φυσική: Δευτέρα 1η (Αίθουσα 2)
    s.add_all([
        TimetableSlot(solution_id=sol.id, lesson_id=l1.id, day_of_week=0,
                      period_id=periods[0].id, classroom_id=room1.id),
        TimetableSlot(solution_id=sol.id, lesson_id=l1.id, day_of_week=1,
                      period_id=periods[1].id, classroom_id=room1.id),
        TimetableSlot(solution_id=sol.id, lesson_id=l2.id, day_of_week=0,
                      period_id=periods[0].id, classroom_id=room2.id),
    ])
    s.commit()

    app = FastAPI()
    app.include_router(exports_router.router, prefix="/api/exports")

    def override_db():
        try:
            yield s
        finally:
            pass

    app.dependency_overrides[get_db] = override_db
    c = TestClient(app)
    c.session = s
    c.sol = sol
    c.t1, c.t2, c.c1, c.c2 = t1, t2, c1, c2
    c.student = student
    yield c
    s.close()


# ------------------------------ baseline (υπάρχοντα) --------------------------

def test_print_single_teacher_contains_their_lessons_only(client):
    res = client.get(f"/api/exports/print?solution_id={client.sol.id}&teacher_id={client.t1.id}")
    assert res.status_code == 200
    assert "Νικολάου Μαρία" in res.text
    assert "ΜΑΘ" in res.text
    assert "ΦΥΣ" not in res.text  # του άλλου καθηγητή


def test_print_requires_exactly_one_filter(client):
    assert client.get(f"/api/exports/print?solution_id={client.sol.id}").status_code == 400
    both = client.get(
        f"/api/exports/print?solution_id={client.sol.id}"
        f"&teacher_id={client.t1.id}&student_id={client.student.id}"
    )
    assert both.status_code == 400


def test_ics_returns_weekly_events(client):
    res = client.get(f"/api/exports/ics?solution_id={client.sol.id}&teacher_id={client.t1.id}")
    assert res.status_code == 200
    assert res.text.count("BEGIN:VEVENT") == 2
    assert "RRULE:FREQ=WEEKLY" in res.text
    # Χωρίς term dates: καμία λήξη/εξαίρεση (η μέχρι-τώρα συμπεριφορά).
    assert "UNTIL=" not in res.text
    assert "EXDATE" not in res.text


def test_ics_with_term_dates_has_anchor_until_and_holiday_exdates(client):
    from backend.models import Term

    s = client.session
    # Σχ. έτος 2026-27: η 25/3/2027 είναι Πέμπτη — βάζουμε μάθημα Πέμπτη
    # για να δούμε EXDATE, και ελέγχουμε άγκυρα/UNTIL στα όρια του σεναρίου.
    term = Term(name="2026-27", is_active=True,
                start_date=__import__("datetime").date(2026, 9, 7),
                end_date=__import__("datetime").date(2027, 5, 28))
    s.add(term)
    s.commit()
    s.refresh(term)
    client.sol.term_id = term.id
    lesson = s.query(Lesson).filter(Lesson.teacher_id == client.t1.id).first()
    period = s.query(Period).first()
    room = s.query(Classroom).first()
    s.add(TimetableSlot(solution_id=client.sol.id, lesson_id=lesson.id,
                        day_of_week=3,  # Πέμπτη
                        period_id=period.id, classroom_id=room.id))
    s.commit()

    res = client.get(f"/api/exports/ics?solution_id={client.sol.id}&teacher_id={client.t1.id}")
    assert res.status_code == 200
    assert "UNTIL=20270528T235959Z" in res.text
    # Άγκυρα: κανένα DTSTART πριν την έναρξη του σεναρίου (7/9/2026).
    import re
    starts = re.findall(r"DTSTART;TZID=Europe/Athens:(\d{8})T", res.text)
    assert starts and all(d >= "20260907" for d in starts)
    # Η 25η Μαρτίου 2027 (Πέμπτη) εξαιρείται από το μάθημα της Πέμπτης.
    assert "EXDATE" in res.text
    assert "20270325T" in res.text


# ------------------------------ μαζική εκτύπωση ------------------------------

def test_bulk_print_all_teachers_one_document_with_page_breaks(client):
    res = client.get(f"/api/exports/print?solution_id={client.sol.id}&all=teachers")
    assert res.status_code == 200
    # Και οι δύο καθηγητές στο ίδιο έγγραφο, ο καθένας σε δική του σελίδα.
    assert "Νικολάου Μαρία" in res.text
    assert "Παπαδόπουλος Γιώργος" in res.text
    assert res.text.count("page-break-after") >= 1


def test_bulk_print_all_classes(client):
    res = client.get(f"/api/exports/print?solution_id={client.sol.id}&all=classes")
    assert res.status_code == 200
    assert "Α1 Λυκείου" in res.text
    assert "Β1 Λυκείου" in res.text


def test_bulk_print_skips_entities_without_slots(client):
    s = client.session
    ghost = Teacher(name="Χωρίς Μαθήματα", short_name="ΧΜ", color="#000")
    s.add(ghost)
    s.commit()
    res = client.get(f"/api/exports/print?solution_id={client.sol.id}&all=teachers")
    assert "Χωρίς Μαθήματα" not in res.text


def test_bulk_print_rejects_combination_with_single_filter(client):
    res = client.get(
        f"/api/exports/print?solution_id={client.sol.id}&all=teachers&teacher_id={client.t1.id}"
    )
    assert res.status_code == 400


def test_bulk_print_unknown_mode_rejected(client):
    res = client.get(f"/api/exports/print?solution_id={client.sol.id}&all=aliens")
    assert res.status_code == 400


# ---------------------------------- XLSX -------------------------------------

def _load_workbook(res):
    from openpyxl import load_workbook

    assert res.status_code == 200, res.text
    assert "spreadsheetml" in res.headers["content-type"]
    return load_workbook(io.BytesIO(res.content))


def test_xlsx_teachers_one_sheet_per_teacher_with_grid(client):
    wb = _load_workbook(client.get(f"/api/exports/xlsx?solution_id={client.sol.id}&mode=teachers"))
    assert len(wb.sheetnames) == 2  # μόνο καθηγητές με slots
    ws = wb[wb.sheetnames[0]]
    header = [cell.value for cell in ws[1]]
    assert "Δευτέρα" in header
    all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "Μαθηματικά" in all_text or "ΜΑΘ" in all_text


def test_xlsx_classes_mode(client):
    wb = _load_workbook(client.get(f"/api/exports/xlsx?solution_id={client.sol.id}&mode=classes"))
    assert len(wb.sheetnames) == 2


def test_xlsx_rooms_mode(client):
    wb = _load_workbook(client.get(f"/api/exports/xlsx?solution_id={client.sol.id}&mode=rooms"))
    assert len(wb.sheetnames) == 2


def test_xlsx_defaults_to_teachers_and_404_on_missing_solution(client):
    wb = _load_workbook(client.get(f"/api/exports/xlsx?solution_id={client.sol.id}"))
    assert len(wb.sheetnames) == 2
    assert client.get("/api/exports/xlsx?solution_id=99999").status_code == 404


def test_xlsx_sheet_titles_are_safe_and_unique(client):
    """Ονόματα με chars που απαγορεύει το Excel σε sheet titles (: / \\ ? * [ ])
    δεν πρέπει να σκάνε το export."""
    s = client.session
    weird = Teacher(name="Κακό/Όνομα: [τεστ]*?", short_name="ΚΟ", color="#000")
    s.add(weird)
    s.commit()
    s.refresh(weird)
    lesson = Lesson(subject_id=client.session.query(Subject).first().id,
                    teacher_id=weird.id, class_id=client.c1.id,
                    periods_per_week=1, duration=1)
    s.add(lesson)
    s.commit()
    s.refresh(lesson)
    period = s.query(Period).first()
    room = s.query(Classroom).first()
    s.add(TimetableSlot(solution_id=client.sol.id, lesson_id=lesson.id,
                        day_of_week=2, period_id=period.id, classroom_id=room.id))
    s.commit()

    wb = _load_workbook(client.get(f"/api/exports/xlsx?solution_id={client.sol.id}&mode=teachers"))
    assert len(wb.sheetnames) == 3
