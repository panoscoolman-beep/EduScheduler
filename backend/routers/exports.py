"""Timetable exports: iCalendar (.ics) feeds, print-friendly HTML and XLSX.

Endpoints (mounted under /api/exports):
    GET /ics?solution_id=&teacher_id=|student_id=   → .ics download
    GET /print?solution_id=&teacher_id=|student_id= → printable HTML grid
    GET /print?solution_id=&all=teachers|classes    → ΟΛΑ τα προγράμματα σε
        ένα έγγραφο, ένα ανά σελίδα (page-break) — «μοίρασε προγράμματα» με 1 κλικ
    GET /xlsx?solution_id=&mode=teachers|classes|rooms → Excel workbook με
        ένα φύλλο ανά καθηγητή/τμήμα/αίθουσα

The frontend opens them with window.open (same-origin auth flow) and the
browser handles the print dialog / download.
"""

from __future__ import annotations

import datetime
from html import escape

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import HTMLResponse, Response
from sqlalchemy.orm import Session, joinedload

from backend.database import get_db
from backend.models import (
    Lesson,
    Period,
    SchoolSettings,
    Student,
    StudentClassEnrollment,
    Teacher,
    Term,
    TimetableSlot,
    TimetableSolution,
)
from backend.services import greek_holidays

router = APIRouter()

_GREEK_DAYS = ["Δευτέρα", "Τρίτη", "Τετάρτη", "Πέμπτη", "Παρασκευή", "Σάββατο", "Κυριακή"]


# ---------------------------------------------------------------------------
# Shared loading / filtering
# ---------------------------------------------------------------------------

def _load_filtered_slots(
    db: Session,
    solution_id: int,
    teacher_id: int | None,
    student_id: int | None,
) -> tuple[TimetableSolution, list[TimetableSlot], str]:
    """Return (solution, placed slots for the filter, filter label)."""
    if (teacher_id is None) == (student_id is None):
        raise HTTPException(
            status_code=400,
            detail="Δώσε ακριβώς ένα από teacher_id ή student_id",
        )

    solution = (
        db.query(TimetableSolution)
        .filter(TimetableSolution.id == solution_id)
        .first()
    )
    if not solution:
        raise HTTPException(status_code=404, detail="Η λύση δεν βρέθηκε")

    query = (
        db.query(TimetableSlot)
        .join(Lesson, TimetableSlot.lesson_id == Lesson.id)
        .filter(
            TimetableSlot.solution_id == solution_id,
            TimetableSlot.is_unplaced == False,  # noqa: E712
        )
        .options(
            joinedload(TimetableSlot.lesson).joinedload(Lesson.subject),
            joinedload(TimetableSlot.lesson).joinedload(Lesson.teacher),
            joinedload(TimetableSlot.lesson).joinedload(Lesson.school_class),
            joinedload(TimetableSlot.classroom),
        )
    )

    if teacher_id is not None:
        teacher = db.query(Teacher).filter(Teacher.id == teacher_id).first()
        if not teacher:
            raise HTTPException(status_code=404, detail="Ο καθηγητής δεν βρέθηκε")
        label = teacher.name
        slots = query.filter(Lesson.teacher_id == teacher_id).all()
    else:
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="Ο μαθητής δεν βρέθηκε")
        label = f"{student.first_name} {student.last_name}"
        class_ids = [
            e.class_id
            for e in db.query(StudentClassEnrollment)
            .filter(StudentClassEnrollment.student_id == student_id)
            .all()
        ]
        slots = query.filter(Lesson.class_id.in_(class_ids)).all() if class_ids else []

    return solution, slots, label


def _periods_by_id(db: Session) -> dict[int, Period]:
    return {p.id: p for p in db.query(Period).all()}


def _get_solution_or_404(db: Session, solution_id: int) -> TimetableSolution:
    solution = (
        db.query(TimetableSolution)
        .filter(TimetableSolution.id == solution_id)
        .first()
    )
    if not solution:
        raise HTTPException(status_code=404, detail="Η λύση δεν βρέθηκε")
    return solution


def _load_all_placed_slots(db: Session, solution_id: int) -> list[TimetableSlot]:
    """Every placed slot of a solution, with the relations the grids need."""
    return (
        db.query(TimetableSlot)
        .join(Lesson, TimetableSlot.lesson_id == Lesson.id)
        .filter(
            TimetableSlot.solution_id == solution_id,
            TimetableSlot.is_unplaced == False,  # noqa: E712
        )
        .options(
            joinedload(TimetableSlot.lesson).joinedload(Lesson.subject),
            joinedload(TimetableSlot.lesson).joinedload(Lesson.teacher),
            joinedload(TimetableSlot.lesson).joinedload(Lesson.school_class),
            joinedload(TimetableSlot.classroom),
        )
        .all()
    )


def _days_for_school(db: Session) -> list[int]:
    """0..N-1 respecting the school's configured week length."""
    settings = db.query(SchoolSettings).first()
    days_per_week = settings.days_per_week if settings else 5
    days_per_week = max(1, min(days_per_week, len(_GREEK_DAYS)))
    return list(range(days_per_week))


def _slot_cell_parts(slot: TimetableSlot, detail_kind: str) -> tuple[str, str, str]:
    """(subject, detail, room) strings for one slot cell.

    detail_kind: τι δείχνουμε κάτω από το μάθημα — 'class' στα προγράμματα
    καθηγητών, 'teacher' στα προγράμματα τμημάτων/μαθητών/αιθουσών."""
    lesson = slot.lesson
    subject = (lesson.subject.short_name or lesson.subject.name) if lesson.subject else "—"
    if detail_kind == "class":
        detail = (
            (lesson.school_class.short_name or lesson.school_class.name)
            if lesson.school_class else ""
        )
    else:
        detail = (lesson.teacher.short_name or lesson.teacher.name) if lesson.teacher else ""
    room = slot.classroom.name if slot.classroom else ""
    return subject, detail, room


def _grid_table_html(
    slots: list[TimetableSlot],
    teaching_periods: list[Period],
    days: list[int],
    detail_kind: str,
) -> str:
    """The weekly grid <table> used by both single and bulk print."""
    grid: dict[tuple[int, int], list[str]] = {}
    for slot in slots:
        if slot.day_of_week is None:
            continue
        subject, detail, room = _slot_cell_parts(slot, detail_kind)
        cell = f"<b>{escape(subject)}</b>"
        if detail:
            cell += f"<br><small>{escape(detail)}</small>"
        if room:
            cell += f"<br><small>🏫 {escape(room)}</small>"
        grid.setdefault((slot.day_of_week, slot.period_id), []).append(cell)

    header_cells = "".join(f"<th>{_GREEK_DAYS[d]}</th>" for d in days)
    rows_html = []
    for p in teaching_periods:
        cells = "".join(
            f"<td>{'<hr>'.join(grid.get((d, p.id), [])) or ''}</td>" for d in days
        )
        rows_html.append(
            f"<tr><th class='time'>{escape(p.start_time)}–{escape(p.end_time)}</th>{cells}</tr>"
        )
    return (
        "<table>"
        f"<thead><tr><th class='time'>Ώρα</th>{header_cells}</tr></thead>"
        f"<tbody>{''.join(rows_html)}</tbody>"
        "</table>"
    )


_PRINT_CSS = """
  body { font-family: 'Segoe UI', Arial, sans-serif; margin: 24px; color: #1a1a2e; }
  h1 { font-size: 20px; margin-bottom: 2px; }
  .sub { color: #666; font-size: 12px; margin-bottom: 16px; }
  table { border-collapse: collapse; width: 100%; }
  th, td { border: 1px solid #999; padding: 6px 8px; text-align: center;
            font-size: 12px; vertical-align: top; }
  thead th { background: #003366; color: white; }
  th.time { background: #eef2f7; white-space: nowrap; }
  td hr { border: none; border-top: 1px dashed #bbb; margin: 4px 0; }
  section.entity { page-break-after: always; }
  section.entity:last-of-type { page-break-after: auto; }
  @media print {
    body { margin: 8mm; }
    .noprint { display: none; }
  }
"""


# ---------------------------------------------------------------------------
# ICS export
# ---------------------------------------------------------------------------

def _next_weekday(base: datetime.date, weekday: int) -> datetime.date:
    """Next date (incl. today) that falls on the given weekday (Mon=0)."""
    return base + datetime.timedelta(days=(weekday - base.weekday()) % 7)


def _ics_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace(",", "\\,").replace(";", "\\;")


@router.get("/ics")
def export_ics(
    solution_id: int,
    teacher_id: int | None = None,
    student_id: int | None = None,
    db: Session = Depends(get_db),
):
    """Weekly-recurring iCalendar feed for one teacher's or student's
    timetable. Import the file into Google Calendar / Outlook / Apple
    Calendar — each lesson becomes a weekly repeating event.

    Αν το σενάριο της λύσης έχει start/end dates, τα events αγκυρώνονται
    στην έναρξη, σταματούν στη λήξη (RRULE UNTIL) και εξαιρούν τις
    ελληνικές αργίες (EXDATE) — αλλιώς τα μαθήματα θα εμφανίζονταν στο
    ημερολόγιο για πάντα, και τον Αύγουστο, και τις αργίες."""
    solution, slots, label = _load_filtered_slots(db, solution_id, teacher_id, student_id)
    periods = _periods_by_id(db)
    today = datetime.date.today()
    stamp = "20260101T000000Z"  # static DTSTAMP: feed is deterministic per solution

    term = (
        db.query(Term).filter(Term.id == solution.term_id).first()
        if solution.term_id is not None else None
    )
    term_start = term.start_date if term and term.start_date else None
    term_end = term.end_date if term and term.end_date else None
    # Άγκυρα: η αρχή του σεναρίου αν υπάρχει (αλλιώς σήμερα, όπως πριν).
    anchor = term_start or today
    holidays = (
        greek_holidays.holidays_in_range(anchor, term_end)
        if term_end and term_end >= anchor else []
    )

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//EduScheduler//Timetable//EL",
        f"X-WR-CALNAME:Πρόγραμμα {_ics_escape(label)}",
        "X-WR-TIMEZONE:Europe/Athens",
        # A VTIMEZONE block so strict parsers (some Outlook versions) accept
        # the TZID=Europe/Athens references below instead of falling back to
        # UTC and shifting every event by the Athens offset.
        "BEGIN:VTIMEZONE",
        "TZID:Europe/Athens",
        "BEGIN:STANDARD",
        "DTSTART:19701025T040000",
        "RRULE:FREQ=YEARLY;BYMONTH=10;BYDAY=-1SU",
        "TZOFFSETFROM:+0300",
        "TZOFFSETTO:+0200",
        "TZNAME:EET",
        "END:STANDARD",
        "BEGIN:DAYLIGHT",
        "DTSTART:19700329T030000",
        "RRULE:FREQ=YEARLY;BYMONTH=3;BYDAY=-1SU",
        "TZOFFSETFROM:+0200",
        "TZOFFSETTO:+0300",
        "TZNAME:EEST",
        "END:DAYLIGHT",
        "END:VTIMEZONE",
    ]

    for slot in slots:
        period = periods.get(slot.period_id)
        if not period or slot.day_of_week is None:
            continue
        start_date = _next_weekday(anchor, slot.day_of_week)
        if term_end and start_date > term_end:
            continue  # σενάριο μικρότερο από εβδομάδα: το μάθημα δεν προλαβαίνει ποτέ
        start_hm = period.start_time.replace(":", "") + "00"
        end_hm = period.end_time.replace(":", "") + "00"
        lesson = slot.lesson
        subject = lesson.subject.name if lesson.subject else "Μάθημα"
        klass = lesson.school_class.name if lesson.school_class else ""
        teacher = lesson.teacher.name if lesson.teacher else ""
        room = slot.classroom.name if slot.classroom else ""

        summary = subject if teacher_id is not None else f"{subject} ({teacher})"
        description = " · ".join(x for x in [klass, teacher, room] if x)

        # RRULE: με DTSTART σε TZID, το UNTIL πρέπει να είναι UTC (RFC 5545).
        rrule = "RRULE:FREQ=WEEKLY"
        if term_end:
            rrule += f";UNTIL={term_end.strftime('%Y%m%d')}T235959Z"

        event = [
            "BEGIN:VEVENT",
            f"UID:eduscheduler-slot-{slot.id}@korifi",
            f"DTSTAMP:{stamp}",
            f"DTSTART;TZID=Europe/Athens:{start_date.strftime('%Y%m%d')}T{start_hm}",
            f"DTEND;TZID=Europe/Athens:{start_date.strftime('%Y%m%d')}T{end_hm}",
            rrule,
        ]
        # EXDATE: αργίες που πέφτουν στη μέρα αυτού του μαθήματος, μετά την
        # άγκυρα — ίδιο TZID+ώρα με το DTSTART ώστε να ταιριάξει το instance.
        exdates = [
            h for h in holidays
            if h.weekday() == slot.day_of_week and h >= start_date
        ]
        for i in range(0, len(exdates), 3):  # ≤3 ανά γραμμή → κάτω από 75 octets
            chunk = exdates[i:i + 3]
            event.append(
                "EXDATE;TZID=Europe/Athens:"
                + ",".join(f"{h.strftime('%Y%m%d')}T{start_hm}" for h in chunk)
            )
        event += [
            f"SUMMARY:{_ics_escape(summary)}",
            f"DESCRIPTION:{_ics_escape(description)}",
            f"LOCATION:{_ics_escape(room)}",
            "END:VEVENT",
        ]
        lines += event

    lines.append("END:VCALENDAR")
    body = "\r\n".join(lines) + "\r\n"
    return Response(
        content=body,
        media_type="text/calendar; charset=utf-8",
        headers={
            "Content-Disposition": (
                f'attachment; filename="timetable_{teacher_id or student_id}.ics"'
            )
        },
    )


# ---------------------------------------------------------------------------
# Print-friendly HTML
# ---------------------------------------------------------------------------

_BULK_MODES = {
    # all=<mode> → (τίτλος, πώς ομαδοποιούμε slots, detail_kind στα κελιά)
    "teachers": ("Προγράμματα Καθηγητών", "teacher", "class"),
    "classes": ("Προγράμματα Τμημάτων", "class", "teacher"),
}


def _bulk_sections(
    db: Session, solution_id: int, mode: str
) -> list[tuple[str, list[TimetableSlot]]]:
    """[(entity label, its slots), ...] sorted by name — μόνο όσοι έχουν slots."""
    slots = _load_all_placed_slots(db, solution_id)
    by_entity: dict[int, list[TimetableSlot]] = {}
    labels: dict[int, str] = {}
    for s in slots:
        lesson = s.lesson
        if mode == "teacher":
            entity = lesson.teacher
        elif mode == "class":
            entity = lesson.school_class
        else:  # room
            entity = s.classroom
        if entity is None:
            continue
        by_entity.setdefault(entity.id, []).append(s)
        labels[entity.id] = entity.name
    return sorted(
        ((labels[eid], entity_slots) for eid, entity_slots in by_entity.items()),
        key=lambda pair: pair[0],
    )


@router.get("/print", response_class=HTMLResponse)
def export_print(
    solution_id: int,
    teacher_id: int | None = None,
    student_id: int | None = None,
    all: str | None = None,
    db: Session = Depends(get_db),
):
    """Printable weekly grid(s). Δύο λειτουργίες:

    - Ένα άτομο: ?teacher_id= ή ?student_id= (όπως πάντα).
    - Μαζικά: ?all=teachers|classes → ΟΛΑ τα προγράμματα σε ένα έγγραφο,
      ένα ανά σελίδα (page-break) — η ροή «μοίρασε προγράμματα» σε 1 κλικ.
    """
    if all is not None:
        if teacher_id is not None or student_id is not None:
            raise HTTPException(
                status_code=400,
                detail="Το all= δεν συνδυάζεται με teacher_id/student_id",
            )
        if all not in _BULK_MODES:
            raise HTTPException(
                status_code=400,
                detail="Το all= δέχεται: teachers ή classes",
            )
        title, group_by, detail_kind = _BULK_MODES[all]
        solution = _get_solution_or_404(db, solution_id)
        periods = sorted(_periods_by_id(db).values(), key=lambda p: p.sort_order)
        teaching_periods = [p for p in periods if not p.is_break]
        days = _days_for_school(db)

        sections_html = []
        for label, entity_slots in _bulk_sections(db, solution_id, group_by):
            table = _grid_table_html(entity_slots, teaching_periods, days, detail_kind)
            sections_html.append(
                "<section class='entity'>"
                f"<h1>📅 Εβδομαδιαίο Πρόγραμμα — {escape(label)}</h1>"
                f"<div class='sub'>Λύση: {escape(solution.name or str(solution.id))} ·"
                f" Εκτυπώθηκε {datetime.date.today().strftime('%d/%m/%Y')} · Φροντιστήριο ΚΟΡΥΦΗ</div>"
                f"{table}</section>"
            )
        body = "".join(sections_html) or "<p>Καμία τοποθετημένη ώρα σε αυτή τη λύση.</p>"
        html = f"""<!DOCTYPE html>
<html lang="el">
<head><meta charset="utf-8"><title>{escape(title)}</title>
<style>{_PRINT_CSS}</style></head>
<body>
{body}
<p class="noprint" style="margin-top:16px">
  <button onclick="window.print()" style="padding:8px 16px">🖨️ Εκτύπωση / Αποθήκευση PDF</button>
</p>
</body>
</html>"""
        return HTMLResponse(content=html)

    solution, slots, label = _load_filtered_slots(db, solution_id, teacher_id, student_id)
    periods = sorted(_periods_by_id(db).values(), key=lambda p: p.sort_order)
    teaching_periods = [p for p in periods if not p.is_break]
    days = _days_for_school(db)
    detail_kind = "class" if teacher_id is not None else "teacher"
    table = _grid_table_html(slots, teaching_periods, days, detail_kind)

    html = f"""<!DOCTYPE html>
<html lang="el">
<head>
<meta charset="utf-8">
<title>Πρόγραμμα — {escape(label)}</title>
<style>{_PRINT_CSS}</style>
</head>
<body>
<h1>📅 Εβδομαδιαίο Πρόγραμμα — {escape(label)}</h1>
<div class="sub">Λύση: {escape(solution.name or str(solution.id))} ·
Εκτυπώθηκε {datetime.date.today().strftime('%d/%m/%Y')} · Φροντιστήριο ΚΟΡΥΦΗ</div>
{table}
<p class="noprint" style="margin-top:16px">
  <button onclick="window.print()" style="padding:8px 16px">🖨️ Εκτύπωση / Αποθήκευση PDF</button>
</p>
</body>
</html>"""
    return HTMLResponse(content=html)


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

_XLSX_MODES = {
    "teachers": ("teacher", "class"),
    "classes": ("class", "teacher"),
    "rooms": ("room", "teacher"),
}

_SHEET_FORBIDDEN = set("[]:*?/\\")


def _safe_sheet_title(name: str, used: set[str]) -> str:
    """Excel: sheet title ≤31 chars, χωρίς []:*?/\\, μοναδικό ανά workbook."""
    cleaned = "".join(c for c in name if c not in _SHEET_FORBIDDEN).strip() or "Φύλλο"
    base = cleaned[:28]
    title = cleaned[:31]
    n = 2
    while title in used:
        title = f"{base}~{n}"
        n += 1
    used.add(title)
    return title


@router.get("/xlsx")
def export_xlsx(
    solution_id: int,
    mode: str = "teachers",
    db: Session = Depends(get_db),
):
    """Excel workbook του ωρολογίου: ένα φύλλο ανά καθηγητή / τμήμα / αίθουσα
    (μόνο όσοι έχουν τοποθετημένες ώρες), το καθένα σε grid Ώρα × Ημέρα."""
    if mode not in _XLSX_MODES:
        raise HTTPException(status_code=400, detail="Το mode= δέχεται: teachers, classes ή rooms")
    group_by, detail_kind = _XLSX_MODES[mode]
    solution = _get_solution_or_404(db, solution_id)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    periods = sorted(_periods_by_id(db).values(), key=lambda p: p.sort_order)
    teaching_periods = [p for p in periods if not p.is_break]
    days = _days_for_school(db)

    wb = Workbook()
    wb.remove(wb.active)  # ξεκινάμε χωρίς το default κενό φύλλο
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="003366")
    wrap = Alignment(wrap_text=True, vertical="top", horizontal="center")
    used_titles: set[str] = set()

    for label, entity_slots in _bulk_sections(db, solution_id, group_by):
        ws = wb.create_sheet(_safe_sheet_title(label, used_titles))
        header = ["Ώρα"] + [_GREEK_DAYS[d] for d in days]
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap

        grid: dict[tuple[int, int], list[str]] = {}
        for slot in entity_slots:
            if slot.day_of_week is None:
                continue
            subject_short, detail, room = _slot_cell_parts(slot, detail_kind)
            lesson = slot.lesson
            subject_full = lesson.subject.name if lesson.subject else subject_short
            text = subject_full
            extras = " · ".join(x for x in [detail, room] if x)
            if extras:
                text += f"\n{extras}"
            grid.setdefault((slot.day_of_week, slot.period_id), []).append(text)

        def _cell_text(parts: list[str]) -> str:
            text = "\n———\n".join(parts)
            # Ένα όνομα που αρχίζει με = / + / @ θα γινόταν ζωντανή φόρμουλα
            # στο Excel — πρόσθεσε απόστροφο ώστε να μείνει κείμενο.
            return "'" + text if text[:1] in ("=", "+", "@") else text

        for p in teaching_periods:
            row = [f"{p.start_time}–{p.end_time}"] + [
                _cell_text(grid.get((d, p.id), [])) for d in days
            ]
            ws.append(row)
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = wrap
        ws.column_dimensions["A"].width = 14
        for i in range(len(days)):
            ws.column_dimensions[chr(ord("B") + i)].width = 26
        ws.freeze_panes = "B2"

    if not wb.sheetnames:  # καμία τοποθετημένη ώρα — δώσε έγκυρο (κενό) αρχείο
        ws = wb.create_sheet("Κενό")
        ws.append(["Καμία τοποθετημένη ώρα σε αυτή τη λύση."])

    import io as _io

    buf = _io.BytesIO()
    wb.save(buf)
    # ASCII-only filename: τα HTTP headers είναι latin-1 — ελληνικό όνομα
    # λύσης στο Content-Disposition σκάει με UnicodeEncodeError.
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="timetable_{solution.id}_{mode}.xlsx"'
            )
        },
    )
